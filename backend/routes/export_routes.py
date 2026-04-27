"""
Smart Store - Export Routes
==============================
Endpoints for exporting data to CSV/Excel.
"""

import io
from flask import Blueprint, request, jsonify, session, Response
from ..db import get_connection, get_dict_cursor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

export_bp = Blueprint("export", __name__)


@export_bp.route("/export/excel", methods=["GET"])
def export_excel():
    """
    Export all owners and products for the logged-in user to a true Excel (.xlsx) file.
    """
    user_id = session.get("user_id") or request.headers.get("X-User-Id")
    if not user_id:
        return jsonify({"success": False, "message": "Unauthorized."}), 401

    try:
        connection = get_connection()
        cursor = get_dict_cursor(connection)

        # Fetch joined data: Owners + Products
        query = """
            SELECT 
                bo.shop_name, 
                bo.owner_name, 
                bo.phone as owner_phone,
                p.id,
                p.display_id,
                p.product_name, 
                p.category, 
                p.price, 
                p.cost_price, 
                p.quantity, 
                p.unit,
                p.is_available
            FROM business_owner bo
            LEFT JOIN products p ON bo.id = p.owner_id
            WHERE bo.user_id = %s
            ORDER BY bo.shop_name ASC, p.display_id ASC
        """
        cursor.execute(query, (user_id,))
        rows = cursor.fetchall()
        
        cursor.close()
        connection.close()

        # Create Excel Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Data"

        # Define Headers
        headers = [
            "Shop Name", "Owner Name", "Owner Phone", 
            "Sequential ID", "System ID (Permanent)", "Product Name", "Category", 
            "Selling Price (Rs)", "Cost Price (Rs)", "Quantity", "Unit", "In Stock"
        ]
        ws.append(headers)

        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add Data Rows
        for row in rows:
            ws.append([
                row.get("shop_name"),
                row.get("owner_name"),
                row.get("owner_phone"),
                row.get("display_id") if row.get("product_name") else "",
                row.get("id") if row.get("product_name") else "",
                row.get("product_name") or "— No Products —",
                row.get("category") or "",
                row.get("price") if row.get("product_name") else "",
                row.get("cost_price") if row.get("product_name") else "",
                row.get("quantity") if row.get("product_name") else "",
                row.get("unit") or "",
                "Yes" if row.get("is_available") else "No" if row.get("product_name") else ""
            ])

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        excel_data = output.getvalue()
        output.close()
        
        filename = f"smart_store_data_{request.args.get('date', 'export')}.xlsx"
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return jsonify({"success": False, "message": f"Excel export failed: {str(e)}"}), 500
